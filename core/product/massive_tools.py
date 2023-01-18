from openpyxl import load_workbook

# Funcion is_none()
def is_none(value):
    data = {}
    if value == None:
        raise Exception('Se ha encontrado un valor nulo donde no deberia estar, corregir el archivo')
    else:
        return value

# Funcion category funca
def is_category(value):
    value = is_none(value)
    value = value.casefold() #Le saco cualquier error humano de poner una mayuscula de mas o de menos
    if value.find(',') >= 1:
        categories = value.split(', ')
        lista = []
        for cat in categories:
            name = cat.capitalize()
            description = f'Categoria de {name.casefold()} - creada automaticamente'
            unit = [name, description]
            lista.append(unit)
        return lista
    else:
        lista = []
        name = value.capitalize()
        description = f'Categoria de {name.casefold()} - creada automaticamente'
        unit = [name, description]
        lista.append(unit)
        return lista


def make_massive_charge(file):
    wb = load_workbook(filename=file)
    work_sheet = wb['Carga Masiva de Productos']
    lista_productos = [row for row in work_sheet.iter_rows(min_row=1, max_col=8) if row[0].value != None]
    del lista_productos[0] # Eliminamos la cabecera
    products = []
    for row in lista_productos:
        unit = []
        for index, cell in enumerate(row):
            if index == 5:
                value = cell.value
                value = is_category(value) # Comprobaremos que sea una categoria y crearemos las categorias necesarias.
                unit.append(value)
            elif index == 4 or index == 6:
                value = cell.value # Los valores podran ser none
                unit.append(value)
            else:
                value = cell.value
                value = is_none(value)
                unit.append(value)
        products.append(unit)
    return products

def check_categories(categories, class_name):
    all_categories = class_name.objects.all()
    lista = []
    for category in categories:
        if class_name.objects.filter(name=category[0]).exists():
            instance = class_name.objects.get(name=category[0])
            lista.append(instance)
        elif category[0] not in all_categories:
            instance = class_name(
                name=category[0],
                description=category[1]
            )
            instance.save()
            #instance = class_name.objects.get(pk=instance.pk)
            lista.append(instance)
        else:
            raise Exception('Error al cargar las categorias')
    return lista



def is_active(value):
    value = value.casefold()
    if value == 'si':
        return True
    elif value == 'no':
        return False
    else:
        raise Exception('Error en columna ¿activo?')